import openpyxl
from openpyxl.chart import BarChart, Reference


def process_wb (filename):

    wb1 = openpyxl.load_workbook(filename)

    sheet = wb1['Sheet1']

    # cell = sheet['a1'] #cell =sheet.cell(1,1)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)  # 2,3// 3,3 // 4,3 = C
        new_value = cell.value * 0.9
        corr = sheet.cell(row, 4)  # reserve columns 4=D
        corr.value = new_value

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb1.save(filename)


file_name = input("it must be a .xlsx file (excel) : ")
process_wb(file_name)